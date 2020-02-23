import json
import requests
import os

# Import `load_workbook` module from `openpyxl`
from openpyxl import load_workbook
from openpyxl import Workbook

# Load and open in the workbook
wb = load_workbook('C:/address_input.xlsx')
wd = Workbook()

# Get a sheet by name 
sheet = wb.get_sheet_by_name('Sheet1')
ws = wd.active

# Get the input address from the Excel sheet, call google API to validate the address, capture the formatted address provided by
# google api and place id in an output excel file
for i in range(1,sheet.max_row):
    ID = sheet.cell(row=i, column=1).value
    address= sheet.cell(row=i, column=2).value
    GOOGLE_MAPS_API_URL = 'https://maps.googleapis.com/maps/api/geocode/json?address='+address+'+CA&key=paste_api_key_here'
    req = requests.get(GOOGLE_MAPS_API_URL)
    res = req.json()
    if req.status_code == 200 and res['status'] == 'OK':
        print("Helo")
        a=0
        for x in res['results']:
            print(a)
            if a ==0:
                ws['A'+str(i)]=ID
                ws['B'+str(i)]=address
                ws['C'+str(i)]=x['formatted_address']
                ws['D'+str(i)]=x['place_id']
                print (x['formatted_address'])
                a = a+1
    else:
        ws['A'+str(i)]=ID
        ws['B'+str(i)]=address
        ws['C'+str(i)]=res['error_message']
        print(res['error_message'])

# Save the ormatted address provided by google api and place id in an output excel file along with input records

wd.save('C:/address_output.xlsx')

print ('Program completed!')
